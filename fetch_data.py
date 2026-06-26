import pandas as pd
import requests
import yfinance as yf
from pathlib import Path
from glob import glob

OUTPUT_DIR = Path(__file__).parent / "data"
OUTPUT_DIR.mkdir(exist_ok=True)

START = "2004-01-01"
END   = "2026-06-25"
FRED_API_KEY = "942d75fa0fda587bf8c9143ded572b70"

# ---------- Yahoo Finance: Gold, VIX, DXY, QQQ, Oil ----------
yf_tickers = {
    "GC=F":     "Gold",
    "^VIX":     "VIX",
    "DX-Y.NYB": "DXY",
    "QQQ":      "QQQ",       # replaced ^GSPC/SP500
    "CL=F":     "Oil_WTI",
    "DJP":      "DJP",       # iPath Bloomberg Commodity ETF (BCOM ex-gold proxy)
}

yf_daily = {}
for ticker, label in yf_tickers.items():
    df = yf.download(ticker, start=START, end=END, progress=False, auto_adjust=True)
    df = df[["Close"]].rename(columns={"Close": label})
    yf_daily[label] = df
    print(f"OK  {label:10s} {len(df)} rows  last: {df.index[-1].date()}")

# ---------- FRED API ----------
# Daily FRED series (can be resampled to both monthly and weekly)
fred_daily_series = {
    "DFII10":  "TIPS_10yr",
    "T10YIE":  "Breakeven",
    "DFF":     "Fed_Rate",
    "BAA10Y":  "Credit_Spread",
}

# Monthly-only FRED series (CPI, M2, EPU — no higher-frequency version)
fred_monthly_series = {
    "CPIAUCSL":   "CPI",
    "M2SL":       "M2",
    "USEPUINDXM": "EPU",
}

def fetch_fred(series_id, label, freq_param=None):
    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}"
        f"&api_key={FRED_API_KEY}"
        f"&observation_start={START}"
        f"&observation_end={END}"
        f"&file_type=json"
    )
    if freq_param:
        url += f"&frequency={freq_param}"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    obs = r.json()["observations"]
    df = pd.DataFrame(obs)[["date", "value"]]
    df.columns = ["DATE", label]
    df["DATE"] = pd.to_datetime(df["DATE"])
    df[label] = pd.to_numeric(df[label], errors="coerce")
    df = df.set_index("DATE").dropna()
    print(f"OK  {label:14s} {len(df)} rows  last: {df.index[-1].date()}")
    return df

fred_daily  = {label: fetch_fred(sid, label) for sid, label in fred_daily_series.items()}
fred_monthly = {label: fetch_fred(sid, label) for sid, label in fred_monthly_series.items()}

# ---------- WGC Gold Demand: CB Net Purchases + ETF Flows (from Excel) ----------
# Looks for any GDT_Tables_Q*_EN.xlsx in the data/ folder (most recent quarter).
# Falls back to cb_demand_wgc.csv (CB only) if no Excel found.

def parse_wgc_excel(path):
    """Extract structural demand quarterly series from WGC Gold Balance sheet.
    Row indices (0-based): Jewellery=11, Technology=14, Bar&Coin=19, ETF=23, CB=24.
    """
    from openpyxl import load_workbook
    wb  = load_workbook(path, read_only=True, data_only=True)
    ws  = wb["Gold Balance"]
    rows = list(ws.iter_rows(values_only=True))

    header    = rows[4]   # year/quarter headers
    jewel_row = rows[11]  # Jewellery
    tech_row  = rows[14]  # Technology
    bc_row    = rows[19]  # Bar & Coin
    etf_row   = rows[23]  # ETFs and Similar Products
    cb_row    = rows[24]  # Central Bank and Other Institutional

    def q_to_date(label):
        q  = int(label[1])
        yr = 2000 + int(label[3:])
        return pd.Timestamp(year=yr, month=q * 3, day=1) + pd.offsets.MonthEnd(0)

    jewel_data, tech_data, bc_data, etf_data, cb_data = {}, {}, {}, {}, {}
    for ci, val in enumerate(header):
        if isinstance(val, str) and len(val) == 5 and val[0] == "Q" and "'" in val:
            date = q_to_date(val)
            for row, d in [(jewel_row, jewel_data), (tech_row, tech_data),
                           (bc_row, bc_data), (etf_row, etf_data), (cb_row, cb_data)]:
                if row[ci] is not None:
                    d[date] = float(row[ci])

    return (pd.Series(jewel_data).sort_index(),
            pd.Series(tech_data).sort_index(),
            pd.Series(bc_data).sort_index(),
            pd.Series(etf_data).sort_index(),
            pd.Series(cb_data).sort_index())


def interp_quarterly(q_series, full_daily_idx):
    """Interpolate quarterly series to daily, then resample to monthly / weekly."""
    combined   = q_series.reindex(q_series.index.union(full_daily_idx)).sort_index()
    daily      = combined.interpolate(method="time")
    return daily, daily.resample("ME").last(), daily.resample("W-FRI").last()


cb_monthly_series    = None; cb_weekly_series    = None; cb_daily_interp    = None
etf_monthly_series   = None; etf_weekly_series   = None; etf_daily_interp   = None
jewel_monthly_series = None; jewel_weekly_series = None; jewel_daily_interp = None
tech_monthly_series  = None; tech_weekly_series  = None; tech_daily_interp  = None
bc_monthly_series    = None; bc_weekly_series    = None; bc_daily_interp    = None

full_daily_idx = pd.date_range(start=START, end=END, freq="D")

wgc_files = sorted(glob(str(OUTPUT_DIR / "GDT_Tables_Q*_EN.xlsx")))
if wgc_files:
    wgc_path = wgc_files[-1]
    print(f"\nWGC Excel: {Path(wgc_path).name}")
    jewel_q, tech_q, bc_q, etf_q, cb_q = parse_wgc_excel(wgc_path)

    _g = globals()
    for q_ser, name, attr in [
        (cb_q,    "CB_Net_Purchases", "cb"),
        (etf_q,   "ETF_Flow",         "etf"),
        (jewel_q, "Jewellery",        "jewel"),
        (tech_q,  "Technology",       "tech"),
        (bc_q,    "Bar_Coin",         "bc"),
    ]:
        d_i, m_s, w_s = interp_quarterly(q_ser, full_daily_idx)
        _g[f"{attr}_daily_interp"]   = d_i
        _g[f"{attr}_monthly_series"] = m_s
        _g[f"{attr}_weekly_series"]  = w_s
        print(f"OK  {name:<20s} {q_ser.dropna().count()} quarters  "
              f"({q_ser.index[0].date()} to {q_ser.index[-1].date()})")
else:
    cb_path = OUTPUT_DIR / "cb_demand_wgc.csv"
    if cb_path.exists():
        cb_q = pd.read_csv(cb_path, index_col=0, parse_dates=True)["CB_Net_Purchases_t"].dropna()
        cb_daily_interp, cb_monthly_series, cb_weekly_series = interp_quarterly(cb_q, full_daily_idx)
        print(f"OK  CB_Net_Purchases (fallback CSV): {cb_monthly_series.dropna().count()} monthly rows")
    else:
        print("SKIP WGC data (no GDT Excel or cb_demand_wgc.csv found in data/)")

# ---------- Build MONTHLY dataset ----------
all_m = {}
for label, df in yf_daily.items():
    s = df[label].squeeze() if hasattr(df[label], "squeeze") else df[label]
    all_m[label] = s.resample("ME").last()
for label, df in fred_daily.items():
    all_m[label] = df[label].resample("ME").last()
for label, df in fred_monthly.items():
    all_m[label] = df[label].resample("ME").last()
for _col, _ser in [("CB_Net_Purchases", cb_monthly_series), ("ETF_Flow", etf_monthly_series),
                   ("Jewellery", jewel_monthly_series), ("Technology", tech_monthly_series),
                   ("Bar_Coin", bc_monthly_series)]:
    if _ser is not None:
        all_m[_col] = _ser

combined_m = pd.DataFrame(all_m)

for col in ["CPI", "M2", "EPU"]:
    if col in combined_m.columns:
        combined_m[col] = combined_m[col].ffill()

combined_m["Gold_Real"] = combined_m["Gold"] / combined_m["CPI"] * 100
combined_m["M2_Growth"] = combined_m["M2"].pct_change(12) * 100

combined_m = combined_m.dropna(subset=["Gold", "TIPS_10yr", "Breakeven", "DXY", "QQQ", "Oil_WTI"])

out_m = OUTPUT_DIR / "gold_factors_monthly.csv"
combined_m.to_csv(out_m)
print(f"\nMonthly: {len(combined_m)} rows  {combined_m.index[0].date()} to {combined_m.index[-1].date()}")
print(f"Saved -> {out_m}")

# ---------- Build WEEKLY dataset ----------
weekly_idx = pd.date_range(start=START, end=END, freq="W-FRI")

all_w = {}
for label, df in yf_daily.items():
    s = df[label].squeeze() if hasattr(df[label], "squeeze") else df[label]
    all_w[label] = s.resample("W-FRI").last().reindex(weekly_idx)

for label, df in fred_daily.items():
    all_w[label] = df[label].resample("W-FRI").last().reindex(weekly_idx)

for label, df in fred_monthly.items():
    # Monthly → weekly: resample to weekly then forward-fill
    s_monthly = df[label].resample("W-FRI").last().reindex(weekly_idx)
    all_w[label] = s_monthly.ffill()

for _col, _ser in [("CB_Net_Purchases", cb_weekly_series), ("ETF_Flow", etf_weekly_series),
                   ("Jewellery", jewel_weekly_series), ("Technology", tech_weekly_series),
                   ("Bar_Coin", bc_weekly_series)]:
    if _ser is not None:
        all_w[_col] = _ser

combined_w = pd.DataFrame(all_w)

for col in ["CPI", "M2", "EPU"]:
    if col in combined_w.columns:
        combined_w[col] = combined_w[col].ffill()

combined_w["Gold_Real"] = combined_w["Gold"] / combined_w["CPI"] * 100
# 52-week M2 growth rate (annualized, comparable to monthly 12m growth)
combined_w["M2_Growth"] = combined_w["M2"].pct_change(52) * 100

combined_w = combined_w.dropna(subset=["Gold", "TIPS_10yr", "Breakeven", "DXY", "QQQ", "Oil_WTI"])

out_w = OUTPUT_DIR / "gold_factors_weekly.csv"
combined_w.to_csv(out_w)
print(f"Weekly:  {len(combined_w)} rows  {combined_w.index[0].date()} to {combined_w.index[-1].date()}")
print(f"Saved -> {out_w}")

print(f"\nLatest monthly values:")
print(combined_m.tail(2).to_string())

# ---------- Build DAILY dataset ----------
# Use Gold trading-day index as reference (yfinance business days)
gold_s = yf_daily["Gold"]["Gold"].squeeze()
gold_s.index = pd.to_datetime(gold_s.index)
trade_idx = gold_s.index  # all Gold trading days

def reindex_to_trading(s, idx):
    """Align any series to the trading day index, forward-filling gaps."""
    s = s.copy()
    s.index = pd.to_datetime(s.index)
    combined = s.reindex(s.index.union(idx)).sort_index().ffill()
    return combined.reindex(idx)

all_d = {}
for label, df in yf_daily.items():
    s = df[label].squeeze() if hasattr(df[label], "squeeze") else df[label]
    s.index = pd.to_datetime(s.index)
    all_d[label] = s.reindex(trade_idx)

for label, df in fred_daily.items():
    all_d[label] = reindex_to_trading(df[label], trade_idx)

for label, df in fred_monthly.items():
    all_d[label] = reindex_to_trading(df[label], trade_idx)

for _col, _ser in [("CB_Net_Purchases", cb_daily_interp), ("ETF_Flow", etf_daily_interp),
                   ("Jewellery", jewel_daily_interp), ("Technology", tech_daily_interp),
                   ("Bar_Coin", bc_daily_interp)]:
    if _ser is not None:
        all_d[_col] = reindex_to_trading(_ser, trade_idx)

combined_d = pd.DataFrame(all_d)
for col in ["CPI", "M2", "EPU"]:
    if col in combined_d.columns:
        combined_d[col] = combined_d[col].ffill()

combined_d["Gold_Real"] = combined_d["Gold"] / combined_d["CPI"] * 100
combined_d["M2_Growth"] = combined_d["M2"].pct_change(252) * 100   # 252-day annualised

combined_d = combined_d.dropna(subset=["Gold", "TIPS_10yr", "DXY", "VIX", "QQQ", "Oil_WTI", "Credit_Spread"])

out_d = OUTPUT_DIR / "gold_factors_daily.csv"
combined_d.to_csv(out_d)
print(f"Daily:   {len(combined_d)} rows  {combined_d.index[0].date()} to {combined_d.index[-1].date()}")
print(f"Saved -> {out_d}")

