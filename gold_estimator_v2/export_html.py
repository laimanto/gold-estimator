"""
Generates Doc/dashboard.html — self-contained gold research dashboard.
"""

import csv, json, math, calendar as _cal
from pathlib import Path
from glob import glob as _glob
from datetime import date as _date, timedelta as _td

MODEL_JSON = Path(__file__).parent / "data" / "model_results.json"
OUT_HTML   = Path(__file__).parent / "Doc" / "dashboard.html"
OUT_HTML.parent.mkdir(exist_ok=True)

with open(MODEL_JSON, encoding="utf-8") as f:
    m = json.load(f)

p1        = m["period1"]
p2        = m["period2"]
p1_cb     = m["period1_cb"]
p1_3m     = m["period1_3m"]
p2_3m     = m["period2_3m"]
p1_3m_w   = m.get("period1_3m_weekly", {})
p2_3m_w   = m.get("period2_3m_weekly", {})
p1_daily  = m.get("period1_daily", {})
p2_daily  = m.get("period2_daily", {})
fv           = m.get("fair_value", {})
cur          = m["current"]
fmeta        = m["factor_meta"]
groups       = m["groups"]
gen          = m.get("generated", "")
stage1_etf   = m.get("stage1_etf", {})
stage1_oil   = m.get("stage1_oil", {})

# ── WGC Structural Demand Breakdown ───────────────────────────────────────
def _parse_wgc_demand_breakdown():
    """Return (data_dict, quarter_labels) from WGC Excel, or (None, [])."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, []
    wgc_files = sorted(_glob(str(Path(__file__).parent / "data" / "GDT_Tables_Q*_EN.xlsx")))
    if not wgc_files:
        return None, []
    wb  = load_workbook(wgc_files[-1], read_only=True, data_only=True)
    ws  = wb["Gold Balance"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[4]
    # Row indices (0-based) confirmed from Gold Balance sheet
    demand_row_idx = {
        "Total Demand":  10,
        "Jewellery":     11,
        "Technology":    14,
        "Bar & Coin":    19,
        "ETF":           23,
        "Central Bank":  24,
    }
    quarter_cols = []
    for ci, val in enumerate(header):
        if isinstance(val, str) and len(val) == 5 and val[0] == "Q" and "'" in val:
            quarter_cols.append((val, ci))
    recent = quarter_cols[-8:]  # last 8 quarters
    result = {}
    for name, ridx in demand_row_idx.items():
        row = rows[ridx]
        series = {}
        for qlabel, ci in recent:
            try:
                v = row[ci]
                if v is not None:
                    series[qlabel] = round(float(v), 1)
            except (TypeError, ValueError):
                pass
        result[name] = series
    return result, [q[0] for q in recent]

_wgc_demand, _wgc_qtrs = _parse_wgc_demand_breakdown()

def _wgc_latest(name):
    vals = list((_wgc_demand or {}).get(name, {}).values())
    return round(vals[-1], 1) if vals else 0.0

_jewel_cur = _wgc_latest("Jewellery")
_tech_cur  = _wgc_latest("Technology")
_bc_cur    = _wgc_latest("Bar & Coin")
_td_cur    = _wgc_latest("Total Demand")

# ── Analyst 3-month consensus forecasts (Q3 2026) ─────────────────────────
_qqq_now = cur.get("QQQ") or 520.0
ANALYST = {
    # ── updated from live sources each session (see feedback_analyst_forecasts.md) ──
    "TIPS_10yr":        2.10,   # FRED DFII10 Jun 26 2026 = 2.17%; 3m consensus ~2.04% (Trading Economics); use 2.10 midpoint
    "Breakeven":        2.17,   # FRED T10YIE Jun 26 2026
    "DXY":            102.5,    # current 101.4; JPM Q3 103-104; firm on Fed-hold (CPI >4%, no cuts); use 102.5
    "M2_Growth":        5.5,    # FRED M2SL trailing trend; no consensus change
    "VIX":             18.0,    # historical mean ~17-19; neutral regime
    "EPU":            225.0,    # recent trailing level (Baker-Bloom-Davis)
    "Credit_Spread":    1.55,   # BAA10Y recent trend; slight widening
    "QQQ":             round(_qqq_now * 1.064, 1),  # Goldman S&P 8,000 = +6.4%
    "Oil_WTI":         72.0,    # Iran war premium unwound; Hormuz recovering; EIA Q3 range $65-78; current ~$70; base $72
    "CB_Net_Purchases": 244.0,  # WGC Q1 2026 actual 244t; FY 850t guidance; Jun 2026 survey: 45% of CBs plan to increase
    "ETF_Flow":         30.0,   # WGC Apr 2026 +45t, May 2026 -16t; Q2 est ~30t (slowing from Q1 +62t)
    "ETF_Residual":      0.0,   # Default = 0 (no speculative/irrational flow assumed)
    "Jewellery":       _jewel_cur,   # WGC Q1 2026 actual; volume under pressure at $3,200+/oz (price elasticity)
    "Technology":      _tech_cur,    # WGC Q1 2026 actual; stable
    "Bar_Coin":        _bc_cur,      # WGC Q1 2026 actual
    "CPI":            335.5,    # reference only
    "DJP":            cur.get("DJP") or 43.0,   # neutral: energy drag from oil decline offset by metals/ag; keep at current
}

# TIPS decompose: nominal ≈ TIPS + Breakeven
_tips_now   = cur.get("TIPS_10yr") or 2.29
_beven_now  = cur.get("Breakeven")  or 2.21
_nom_now    = round(_tips_now + _beven_now, 2)
_tips_fcast = ANALYST["TIPS_10yr"]
_beven_fcast= ANALYST["Breakeven"]
_nom_fcast  = round(_tips_fcast + _beven_fcast, 2)

# ── Ticker / source info per factor ───────────────────────────────────────
TICKER_INFO = {
    "D_TIPS_10yr": {
        "futu": None,
        "description": "Unit: Δ percentage points (pp). Frequency: Daily.",
        "url": "https://fred.stlouisfed.org/series/DFII10",
        "url_label": "FRED: DFII10",
    },
    "D_ln_DXY": {
        "futu": "DXY",
        "description": "Unit: Δ% log-return. Frequency: Daily.",
        "url": "https://finance.yahoo.com/quote/DX-Y.NYB/",
        "url_label": "Yahoo Finance: DX-Y.NYB",
    },
    "M2_Growth": {
        "futu": None,
        "description": "Unit: % YoY (level). Frequency: Monthly (~4-week lag after month-end; quarterly revision).",
        "url": "https://fred.stlouisfed.org/series/M2SL",
        "url_label": "FRED: M2SL",
    },
    "D_ln_VIX": {
        "futu": ".VIX",
        "description": "Unit: Δ% log-return. Frequency: Daily.",
        "url": "https://finance.yahoo.com/quote/%5EVIX/",
        "url_label": "Yahoo Finance: ^VIX",
    },
    "D_ln_EPU": {
        "futu": None,
        "description": "Unit: Δ% log-return. Frequency: Monthly (~1-month lag). Academic index — no live market ticker.",
        "url": "https://fred.stlouisfed.org/series/USEPUINDXM",
        "url_label": "FRED: USEPUINDXM",
    },
    "D_Credit_Spread": {
        "futu": None,
        "description": "Unit: Δ percentage points (pp). Frequency: Daily.",
        "url": "https://fred.stlouisfed.org/series/BAA10Y",
        "url_label": "FRED: BAA10Y",
    },
    "D_ln_QQQ": {
        "futu": "QQQ",
        "description": "Unit: Δ% log-return. Frequency: Daily.",
        "url": "https://finance.yahoo.com/quote/QQQ/",
        "url_label": "Yahoo Finance: QQQ",
    },
    "D_ln_Oil": {
        "futu": "CL.main",
        "description": "Unit: Δ% log-return. Frequency: Daily. QE Era only — raw oil not used in Dollar Era model.",
        "url": "https://finance.yahoo.com/quote/CL%3DF/",
        "url_label": "Yahoo Finance: CL=F",
    },
    "Oil_Residual": {
        "futu": "CL.main",
        "description": (
            "Unit: $/bbl input → Δ% residual used in OLS. Frequency: Daily. "
            "Dollar Era: Stage 1b macro-stripped residual (actual Δln_Oil minus TIPS/DXY/M2-predicted). "
            "QE Era: raw Δln_Oil log-return (no orthogonalization)."
        ),
        "url": "https://finance.yahoo.com/quote/CL%3DF/",
        "url_label": "Yahoo Finance: CL=F",
    },
    "CB_Net_Purchases": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter. Frequency: Quarterly (~6–8 weeks after quarter end: Q1→May, Q2→Aug, Q3→Nov, Q4→Feb). "
            "Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "ETF_Flow": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter. Frequency: Quarterly (~6–8 weeks after quarter end). "
            "Reference only — active OLS factor is ETF_Residual (macro-stripped). "
            "Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "ETF_Residual": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter (enter expected total ETF inflow). "
            "Dashboard auto-strips the macro-justified portion — enter raw total ETF flow forecast. "
            "Frequency: Quarterly. Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "Jewellery": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter. Frequency: Quarterly (~6–8 weeks after quarter end). "
            "Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "Technology": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter. Frequency: Quarterly (~6–8 weeks after quarter end). "
            "Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "Bar_Coin": {
        "futu": None,
        "description": (
            "Unit: tonnes/quarter. Frequency: Quarterly (~6–8 weeks after quarter end). "
            "Free WGC registration required — download GDT_Tables_Q?xx_EN.xlsx and place in data/ folder."
        ),
        "url": "https://www.gold.org/goldhub/research/gold-demand-trends",
        "url_label": "WGC: Gold Demand Trends",
    },
    "D_ln_BCOM_exgold": {
        "futu": "DJP",
        "description": (
            "Unit: $ DJP price input → Δ% ex-gold log-return in OLS. Frequency: Daily. "
            "Dashboard auto-strips gold's ~12.7% index weight: (Δln_DJP − 0.127×Δln_Gold)/0.873."
        ),
        "url": "https://finance.yahoo.com/quote/DJP/",
        "url_label": "Yahoo Finance: DJP",
    },
}

def ticker_badge(col):
    futu = TICKER_INFO.get(col, {}).get("futu")
    if not futu: return ""
    return (f'<span style="display:inline-block;margin-left:6px;background:#2a1e00;'
            f'color:#c9a84c;font-size:10px;border-radius:3px;padding:1px 5px;'
            f'font-family:monospace">{futu}</span>')

def info_expand_html(col, prefix):
    """Returns (inline_button_html, panel_html) for the expandable +info block."""
    info = TICKER_INFO.get(col)
    if not info: return "", ""
    futu = info.get("futu")
    futu_line = (f'<span style="color:#c9a84c;font-family:monospace">{futu}</span>'
                 f' (Futu ticker) &nbsp;|&nbsp; ') if futu else ""
    uid = f"info_{prefix}_{col}"
    btn = (f'<button onclick="toggleEl(\'{uid}\')" style="font-size:10px;background:#1a1d27;'
           f'border:1px solid #333;color:#555;border-radius:3px;padding:0px 4px;'
           f'cursor:pointer;line-height:1.4;margin-left:5px">+info</button>')
    panel = (f'<div id="{uid}" style="display:none;font-size:11px;color:#888;margin-top:5px;'
             f'max-width:260px;line-height:1.45;padding:6px 8px;background:#12141e;border-radius:5px">'
             f'{info["description"]}<br>'
             f'<div style="margin-top:4px">{futu_line}'
             f'<a href="{info["url"]}" target="_blank" style="color:#7abfff">{info["url_label"]}</a>'
             f'</div></div>')
    return btn, panel

# ── Helpers ────────────────────────────────────────────────────────────────
def raw_key(col):  return fmeta[col]["raw_col"]
def cur_val(col):  return cur.get(raw_key(col))

def fmt_display(col, val=None):
    if val is None: val = cur_val(col)
    if val is None: return "N/A"
    rk = raw_key(col)
    if rk == "EPU":              return f"{val:.0f}"
    if rk == "CB_Net_Purchases": return f"{val:.1f}"
    return f"{val:.2f}"

def input_num(col):
    rk  = raw_key(col)
    val = ANALYST.get(rk, cur_val(col))
    if val is None: return "0"
    if rk == "EPU": return f"{val:.0f}"
    return f"{val:.2f}"

def is_changed(col):
    rk  = raw_key(col)
    inp = float(input_num(col))
    c   = cur_val(col) or 0
    thr = 0.5 if rk == "EPU" else 0.005
    return abs(inp - c) > thr

def fmt_beta(v):
    """Standardized coefficient: +0.43 or −0.09, directly comparable."""
    return f"{v:+.2f}" if v is not None else "—"

def fmt_rp(v):
    """Normalized partial R²% — share of explained variance, sums to 100%."""
    if v is None or v == 0: return "—"
    if v < 0.05: return "<0.1%"
    return f"{v:.1f}%"

def sig_badge(s):
    c = {"***": "#4caf82", "**": "#a8d4b8", "*": "#d4c882"}.get(s, "#555")
    return f'<span style="color:{c};font-size:11px">{s if s else "n.s."}</span>'

def shift_arrow(b1, b2):
    if b1 is None or b2 is None: return ""
    if (b1 > 0) != (b2 > 0):
        return '<span style="color:#e05c5c" title="Sign reversed">&#8644;</span>'
    if abs(b2) > abs(b1) * 1.2:
        return '<span style="color:#e05c5c" title="Strengthened">&#8593;</span>'
    if abs(b2) < abs(b1) * 0.8:
        return '<span style="color:#4caf82" title="Weakened">&#8595;</span>'
    return '<span style="color:#888">&#8594;</span>'

# ── Computed stats ─────────────────────────────────────────────────────────
# Use 3m weekly HAC model RMSE directly (already quarterly scale — no √3 needed)
rmse3m   = p2_3m_w.get("rmse") or p2["rmse"] * math.sqrt(3)
sd_dol   = round(cur["Gold_Nominal"] * (math.exp(rmse3m) - 1))
ci_dol   = round(cur["Gold_Nominal"] * (math.exp(1.96 * rmse3m) - 1))
p1_3m_w_r2  = p1_3m_w.get("r2")
p2_3m_w_r2  = p2_3m_w.get("r2")
p1_3m_w_n   = p1_3m_w.get("n", "?")
p2_3m_w_n   = p2_3m_w.get("n", "?")
p1_daily_r2 = p1_daily.get("r2")
p2_daily_r2 = p2_daily.get("r2")

fv_pct3   = fv.get("trailing_3m_pct", 0)
fv_pct6   = fv.get("trailing_6m_pct", 0)
fv_z      = fv.get("z_score_3m", 0)
fv_desc   = fv.get("description", "")
fv_price  = fv.get("fair_value_price", None)        # model-implied fair value price
fv_from_cur = fv.get("fair_value_pct_from_current", 0)  # % from current to fair value

def fv_color(pct):
    if pct > 5:   return "#e05c5c"   # above model = stretched (red)
    if pct < -5:  return "#4caf82"   # below model = lagging (green)
    return "#888"

# ── Comparison table rows ──────────────────────────────────────────────────

def group_header(name, colspan=8):
    return (f'<tr class="fa-group-hdr" style="background:#12141e"><td colspan="{colspan}" '
            f'style="padding:10px 8px 6px;color:#c9a84c;font-weight:700;'
            f'font-size:12px;letter-spacing:0.5px;text-transform:uppercase">{name}</td></tr>')

def _combined_row(col, category, p1w_factors, p2w_factors, p1cb_factors):
    """Combined Factor Analysis + Estimator row (8 columns)."""
    p1f = p1w_factors.get(col)
    p2f = p2w_factors.get(col)
    _p1f_src = p1f  # source used for rp1_str guard (may be overridden for CB)
    if col == "CB_Net_Purchases" and not p1f:
        p1f_cb   = p1cb_factors.get(col)
        _p1f_src = p1f_cb
        b1  = p1f_cb.get("std_beta")       if p1f_cb else None
        s1  = p1f_cb.get("sig", "")        if p1f_cb else ""
        rp1 = p1f_cb.get("partial_r2_pct") if p1f_cb else None
    else:
        b1  = p1f.get("std_beta")       if p1f else None
        s1  = p1f.get("sig", "")        if p1f else ""
        rp1 = p1f.get("partial_r2_pct") if p1f else None

    b2  = p2f.get("std_beta")       if p2f else None
    s2  = p2f.get("sig", "")        if p2f else ""
    rp2 = p2f.get("partial_r2_pct") if p2f else None
    # Factor in model but rounds to 0.0 → show <0.1% not blank (applies to both eras)
    rp1_str = ("<0.1%" if (_p1f_src is not None and (rp1 is None or rp1 == 0)) else fmt_rp(rp1))
    rp2_str = ("<0.1%" if (p2f is not None and (rp2 is None or rp2 == 0)) else fmt_rp(rp2))

    info    = fmeta.get(col, {})
    label   = info.get("label", col)
    unit    = info.get("unit", "")
    cb_note = ' <small style="color:#666">(2016–22 monthly)</small>' if col == "CB_Net_Purchases" else ""

    # Estimator inputs
    cv       = fmt_display(col)
    inp      = input_num(col)
    cur_num  = cur_val(col) or 0
    changed  = is_changed(col)
    shade    = "#2a1e00" if changed else "#0f1117"
    border   = "#c9a84c44" if changed else "#333"

    da = (f' data-beta2="{b2 if b2 is not None else 0}"'
          f' data-rp2="{rp2 if rp2 is not None else 0}"'
          f' data-beta1="{b1 if b1 is not None else 0}"'
          f' data-rp1="{rp1 if rp1 is not None else 0}"')

    info_btn, info_panel = info_expand_html(col, "est")
    cat_tag = f'<div style="font-size:10px;color:#888;margin-top:1px">{category}{info_btn}</div>'

    row = (
        f'<tr class="fa-row"{da}>'
        f'<td><b>{label}</b>{cb_note}{ticker_badge(col)}'
        f'{cat_tag}{info_panel}</td>'
        f'<td style="text-align:center;color:#c9a84c;font-size:12px">{cv}</td>'
        f'<td style="text-align:center">{fmt_beta(b1)}<br>{sig_badge(s1)}</td>'
        f'<td style="text-align:center;color:#7abfff">{rp1_str}</td>'
        f'<td style="text-align:center">{fmt_beta(b2)}<br>{sig_badge(s2)}</td>'
        f'<td style="text-align:center;color:#ff9f4a">{rp2_str}</td>'
        f'<td style="text-align:center">'
        f'<input type="number" id="fi_{col}" data-current="{cur_num}" value="{inp}" step="any"'
        f' style="width:90px;background:{shade};color:#c9a84c;border:1px solid {border};'
        f'border-radius:4px;padding:4px 6px;text-align:center;font-size:12px;font-weight:400;transition:background 0.2s"'
        f' oninput="onInput(this)"></td>'
        f'<td style="text-align:center" id="fi_impact_{col}">—</td>'
        f'</tr>'
    )

    if col == "D_TIPS_10yr":
        row += (
            f'<tr><td colspan="8" style="padding:2px 0 8px">'
            f'<button onclick="toggleEl(\'tips_expand_wrap\')" '
            f'style="font-size:10px;background:#1a1d27;border:1px solid #2a2d3a;color:#777;'
            f'border-radius:3px;padding:2px 8px;cursor:pointer;margin-left:8px">'
            f'+ Decompose: Nominal Rate &#8722; Inflation Expectation</button>'
            f'<div id="tips_expand_wrap" style="display:none;margin:8px;'
            f'background:#12141e;border-radius:6px;padding:12px 14px;border:1px solid #2a2d3a">'
            f'<div style="color:#888;font-size:11px;margin-bottom:10px">'
            f'Real Rate = 10yr Nominal Rate &#8722; Market Inflation Expectation (Fisher equation)'
            f'</div>'
            f'<div style="display:flex;gap:16px;align-items:flex-end;flex-wrap:wrap">'
            f'<div><label style="color:#666;font-size:11px">10yr Nominal Rate (%)<br>'
            f'<a href="https://finance.yahoo.com/quote/%5ETNX/" target="_blank" style="color:#7abfff">Yahoo: ^TNX</a>'
            f' / <a href="https://fred.stlouisfed.org/series/GS10" target="_blank" style="color:#7abfff">FRED: GS10</a>'
            f'</label><br>'
            f'<input type="number" id="fi_nom_rate" value="{_nom_fcast:.2f}" step="0.01"'
            f' style="width:80px;background:#0f1117;color:#b8bcc8;border:1px solid #333;'
            f'border-radius:4px;padding:4px 6px;text-align:center" oninput="updateTipsFromParts()">'
            f'<br><small style="color:#555">current: {_nom_now:.2f}%</small></div>'
            f'<div style="color:#c9a84c;font-size:20px;padding-bottom:18px">&#8722;</div>'
            f'<div><label style="color:#666;font-size:11px">10yr Inflation Expectation (%)<br>'
            f'<a href="https://fred.stlouisfed.org/series/T10YIE" target="_blank" style="color:#7abfff">FRED: T10YIE (Breakeven)</a>'
            f'</label><br>'
            f'<input type="number" id="fi_inf_exp" value="{_beven_fcast:.2f}" step="0.01"'
            f' style="width:80px;background:#0f1117;color:#b8bcc8;border:1px solid #333;'
            f'border-radius:4px;padding:4px 6px;text-align:center" oninput="updateTipsFromParts()">'
            f'<br><small style="color:#555">current: {_beven_now:.2f}%</small></div>'
            f'<div style="color:#c9a84c;font-size:20px;padding-bottom:18px">=</div>'
            f'<div><label style="color:#666;font-size:11px">Computed TIPS (%)</label><br>'
            f'<div id="tips_computed_val" style="font-size:20px;font-weight:700;color:#c9a84c;padding:4px 0">'
            f'{_tips_fcast:.2f}%</div>'
            f'<small style="color:#555">auto-fills TIPS input above</small></div>'
            f'</div></div>'
            f'</td></tr>'
        )
    return row

def _combined_struct_rows(p1w_factors, p2w_factors, p1cb_factors):
    """Structural Demand block: Total summary row + 5 OLS sub-components."""
    rows = []
    struct_cols = ["CB_Net_Purchases", "ETF_Residual", "Jewellery", "Technology", "Bar_Coin"]
    def _rp(factors, col):
        f = factors.get(col) or p1cb_factors.get(col) or {}
        return f.get("partial_r2_pct") or 0
    tot_rp1 = sum(_rp(p1w_factors, c) for c in struct_cols)
    tot_rp2 = sum(_rp(p2w_factors, c) for c in struct_cols)

    rows.append(
        f'<tr class="fa-struct-hdr" style="background:#161920">'
        f'<td><b>Total Structural Demand</b>'
        f'<div style="font-size:10px;color:#888;margin-top:1px">Structural Demand</div></td>'
        f'<td style="text-align:center;color:#c9a84c;font-size:12px" id="fi_cur_struct_total">{_td_cur:,.0f}t</td>'
        f'<td style="text-align:center;color:#888">—</td>'
        f'<td style="text-align:center;color:#7abfff">{fmt_rp(tot_rp1) if tot_rp1 else "—"}</td>'
        f'<td style="text-align:center;color:#888">—</td>'
        f'<td style="text-align:center;color:#ff9f4a">{fmt_rp(tot_rp2) if tot_rp2 else "—"}</td>'
        f'<td style="text-align:center">'
        f'<div id="fi_struct_total" style="color:#c9a84c;font-size:13px;font-weight:600">—</div></td>'
        f'<td style="text-align:center;color:#888;font-size:11px">sum</td>'
        f'</tr>'
    )
    for col in struct_cols:
        if col in fmeta:
            rows.append(_combined_row(col, "Structural Demand", p1w_factors, p2w_factors, p1cb_factors))
    return "\n".join(rows)

# Build combined estimator table (Factor Analysis + Estimator merged)
p1w_f  = p1_3m_w.get("factors", {})
p2w_f  = p2_3m_w.get("factors", {})
p1cb_f = p1_cb.get("factors", {})

# Map each group name to a category label for inline display
_group_to_cat = {gname: gname for gname, _ in groups}

rows_html = []
for gname, gcols in groups:
    rows_html.append(group_header(gname))
    if gname == "Structural Demand":
        rows_html.append(_combined_struct_rows(p1w_f, p2w_f, p1cb_f))
    else:
        for col in gcols:
            if col in fmeta:
                rows_html.append(_combined_row(col, gname, p1w_f, p2w_f, p1cb_f))

_p1w_adj = p1_3m_w.get("r2_adj")
_p2w_adj = p2_3m_w.get("r2_adj")
rows_html.append(
    f'<tr class="fa-footer" style="background:#12141e;border-top:2px solid #2a2d3a">'
    f'<td style="color:#b8bcc8;font-weight:700;padding:10px 8px">'
    f'ALL FACTORS — 3m Weekly R²<br>'
    f'<small style="color:#555;font-weight:normal">Weight% sums to 100% of explained variance</small></td>'
    f'<td></td>'
    f'<td colspan="2" style="text-align:center;color:#7abfff;font-weight:700">'
    f'{f"{p1_3m_w_r2:.3f} = {p1_3m_w_r2*100:.0f}%" if p1_3m_w_r2 else "—"}<br>'
    f'<small style="font-weight:normal;color:#5599cc">N={p1_3m_w_n}w &nbsp; Adj={f"{_p1w_adj:.3f}" if _p1w_adj else "—"}</small></td>'
    f'<td colspan="2" style="text-align:center;color:#ff9f4a;font-weight:700">'
    f'{f"{p2_3m_w_r2:.3f} = {p2_3m_w_r2*100:.0f}%" if p2_3m_w_r2 else "—"}<br>'
    f'<small style="font-weight:normal;color:#cc7733">N={p2_3m_w_n}w &nbsp; Adj={f"{_p2w_adj:.3f}" if _p2w_adj else "—"}</small></td>'
    f'<td colspan="2" style="color:#555;font-size:11px;padding-left:6px">'
    f'β* = std. coeff (1 SD in X → β* SD in gold). Weight% = share of explained variance.<br>'
    f'Structural demand% = WGC volume share of total demand (≠ OLS weight).</td>'
    f'</tr>'
)
combined_html = "\n".join(rows_html)

# ── Last-update display (use file mtime for time precision) ───────────────
from datetime import datetime as _dt, timezone as _tz
import os as _os
try:
    _json_path  = Path(__file__).parent / "data" / "model_results.json"
    _mtime      = _os.path.getmtime(str(_json_path))
    _gen_dt     = _dt.fromtimestamp(_mtime, tz=_tz.utc)
    gen_display = _gen_dt.strftime("%d %b %Y %H:%M:%S UTC")
except Exception:
    try:
        _gen_dt     = _dt.strptime(gen[:10], "%Y-%m-%d")
        gen_display = _gen_dt.strftime("%d %b %Y UTC")
    except Exception:
        gen_display = gen

# ── JS data — uses 3m weekly HAC model throughout ─────────────────────────
# Weekly HAC coef is quarterly-scale for level factors (no ×3 needed in JS)
def get_coef_w(col):
    f = p2_3m_w.get("factors", {}).get(col) or {}
    return f.get("coef", 0.0)

js_factors = {
    col: {
        "coef":       get_coef_w(col),
        "input_type": info["input_type"],
        "current":    cur.get(info["raw_col"]),
        "is_level":   (info["input_type"] == "level"),
    }
    for col, info in fmeta.items()
}

js_model = {
    "intercept":        p2_3m_w.get("intercept", p2["intercept"]),
    "rmse":             rmse3m,          # already 3m scale
    "currentLnGoldNom": cur["ln_Gold_Nominal"],
    "currentGoldNom":   cur["Gold_Nominal"],
    "fvPrice":          fv_price,        # model fair value price (current period residual)
    "factors":          js_factors,
    "stage1_etf":       stage1_etf,      # Stage 1a coefs: ETF_Flow ~ TIPS + DXY + M2
    "stage1_oil":       stage1_oil,      # Stage 1b coefs: D_ln_Oil ~ TIPS + DXY + M2
}
js_json = json.dumps(js_model, indent=2)


# ── Methodology table ──────────────────────────────────────────────────────
# method_rows: (group, factor_name, unit, description)
# group=None → no group header (dependent variable row)
method_rows = [
    (None, "Gold (dependent)", "Δln(Gold_Nominal)",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/GC=F/' style='color:#7abfff'>Yahoo Finance: GC=F</a>.<br>"
     "Monthly % return in nominal gold price. "
     "No CPI deflation needed — inflation is captured through TIPS "
     "(nominal rate = TIPS + breakeven inflation), so deflating would be redundant. "
     "Δln(Gold) gives a stationary proportional return series suitable for OLS."),
    ("Rates & Inflation", "Real Rate (TIPS)", "Δ percentage points",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://fred.stlouisfed.org/series/DFII10' style='color:#7abfff'>FRED: DFII10</a>.<br>"
     "Monthly change in the 10yr TIPS yield = the new information that triggers reallocation. "
     "Breakeven (inflation expectation) is shown as a decomposition only — it is collinear "
     "with TIPS since nominal rate = TIPS + breakeven, so including both would introduce "
     "perfect multicollinearity. TIPS directly captures the combined real-rate effect."),
    ("US Dollar & Money Supply", "US Dollar (DXY)", "Δ% log-return",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/DX-Y.NYB/' style='color:#7abfff'>Yahoo Finance: DX-Y.NYB</a>.<br>"
     "Log-diff captures proportional impact: 1% dollar move at DXY=90 and DXY=110 "
     "both represent the same purchasing-power shift for global gold buyers."),
    ("US Dollar & Money Supply", "Money Supply (M2)", "Level % YoY",
     "<b>Frequency:</b> Monthly (~4-week lag after month-end; quarterly revision). "
     "<b>Source:</b> <a href='https://fred.stlouisfed.org/series/M2SL' style='color:#7abfff'>FRED: M2SL</a>.<br>"
     "An ongoing M2 growth rate of 5% creates a persistent tailwind each month — "
     "differencing would capture only acceleration/deceleration, missing the continuous effect."),
    ("Risk & Uncertainty", "Market Fear (VIX)", "Δ% log-return",
     "<b>Frequency:</b> Daily (real-time during US market hours). "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/%5EVIX/' style='color:#7abfff'>Yahoo Finance: ^VIX</a>.<br>"
     "Proportional VIX moves drive equivalent safe-haven flows. VIX is log-normal (Whaley 2009)."),
    ("Risk & Uncertainty", "Policy Uncertainty (EPU)", "Δ% log-return",
     "<b>Frequency:</b> Monthly (~1-month lag; academic index, no live market ticker). "
     "<b>Source:</b> <a href='https://fred.stlouisfed.org/series/USEPUINDXM' style='color:#7abfff'>FRED: USEPUINDXM</a> (Baker, Bloom &amp; Davis).<br>"
     "Log-diff normalises extreme spikes (Baker/Bloom/Davis 2016 use log-EPU in their regressions)."),
    ("Risk & Uncertainty", "Credit / Default Risk", "Δ percentage points",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://fred.stlouisfed.org/series/BAA10Y' style='color:#7abfff'>FRED: BAA10Y</a> (Moody's Baa minus 10yr Treasury).<br>"
     "A +0.5pp widening signals the same incremental default stress regardless of the absolute spread level."),
    ("Financial Markets", "Stock Market (QQQ)", "Δ% log-return",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/QQQ/' style='color:#7abfff'>Yahoo Finance: QQQ</a>.<br>"
     "Monthly % return — risk-off rotation into gold responds to the magnitude of equity loss, "
     "not the index level. QQQ used for higher beta exposure vs S&P 500."),
    ("Financial Markets", "Commodity Index (ex-Gold)", "Δ% log-return",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/DJP/' style='color:#7abfff'>Yahoo Finance: DJP</a> (iPath Bloomberg Commodity ETF).<br>"
     "DJP log-return mathematically stripped of gold's ~12.7% index weight: "
     "(Δln_DJP − 0.127 × Δln_Gold) / 0.873. Removes the circularity where BCOM "
     "partially explains gold because gold IS in BCOM. The residual captures commodity supercycle "
     "dynamics — energy, industrial metals, agriculture — that independently co-move with gold "
     "through inflation expectations and risk-on/off channels. "
     "Partial R²: 3.7%** (QE Era) / 12.0%*** (Dollar Era)."),
    ("Financial Markets", "Oil (Commodity-Adj.)", "Δ% residual log-return",
     "<b>Frequency:</b> Daily. "
     "<b>Source:</b> <a href='https://finance.yahoo.com/quote/CL%3DF/' style='color:#7abfff'>Yahoo Finance: CL=F</a> (WTI front-month futures).<br>"
     "Stage 1b orthogonalization: D_ln_Oil regressed on TIPS, DXY, and M2 to strip the "
     "inflation-proxy component already captured by TIPS. "
     "Oil_Residual = actual Δln(Oil) − macro-predicted = pure localized commodity dynamics "
     "(OPEC+ quota shifts, refinery bottlenecks, strategic reserve releases). "
     "Negligible residual weight (~0.6%) confirms oil's systemic signal is fully absorbed by TIPS."),
    ("Structural Demand", "CB Net Purchases", "tonnes/quarter (t/qtr)",
     "<b>Frequency:</b> Quarterly (~6–8 weeks after quarter end: Q1→May, Q2→Aug, Q3→Nov, Q4→Feb). "
     "<b>Source:</b> <a href='https://www.gold.org/goldhub/research/gold-demand-trends' style='color:#7abfff'>WGC: Gold Demand Trends</a> "
     "(GDT_Tables_Q?xx_EN.xlsx, Gold Balance sheet — free registration required).<br>"
     "Strategic sovereign reserve accumulation (de-dollarization, sanction-proofing). "
     "Price-inelastic and independent of short-term macro conditions. "
     "No differencing needed — level directly measures ongoing accumulation pace."),
    ("Structural Demand", "ETF Flow (Macro-Adj.)", "tonnes/quarter (t/qtr, raw input)",
     "<b>Frequency:</b> Quarterly (~6–8 weeks after quarter end). "
     "<b>Source:</b> <a href='https://www.gold.org/goldhub/research/gold-demand-trends' style='color:#7abfff'>WGC: Gold Demand Trends</a> "
     "(GDT_Tables_Q?xx_EN.xlsx — free registration required).<br>"
     "Stage 1a orthogonalization: ETF_Flow regressed on TIPS, DXY, and M2 to remove "
     "the macro-driven component (algorithmic reallocation triggered by rate moves). "
     "ETF_Residual = actual flow − macro-predicted = pure speculative / retail sentiment foam. "
     "Fitted separately for each era using era-specific weekly data. "
     "User enters expected total ETF inflow; the orthogonalized β accounts for the macro-justified portion."),
    ("Structural Demand", "Jewellery Demand", "tonnes/quarter (t/qtr)",
     "<b>Frequency:</b> Quarterly (~6–8 weeks after quarter end). "
     "<b>Source:</b> <a href='https://www.gold.org/goldhub/research/gold-demand-trends' style='color:#7abfff'>WGC: Gold Demand Trends</a> "
     "(GDT_Tables_Q?xx_EN.xlsx, Gold Balance sheet row 12 — free registration required).<br>"
     "WGC quarterly consumer jewellery purchases. Driven by income, cultural demand cycles "
     "(Indian wedding season, Lunar New Year), and gold price elasticity — high prices suppress "
     "jewellery buying. Interpolated from quarterly to weekly frequency."),
    ("Structural Demand", "Technology Demand", "tonnes/quarter (t/qtr)",
     "<b>Frequency:</b> Quarterly (~6–8 weeks after quarter end). "
     "<b>Source:</b> <a href='https://www.gold.org/goldhub/research/gold-demand-trends' style='color:#7abfff'>WGC: Gold Demand Trends</a> "
     "(GDT_Tables_Q?xx_EN.xlsx, Gold Balance sheet row 15 — free registration required).<br>"
     "WGC quarterly industrial demand (semiconductor bonding, medical devices). "
     "Relatively stable and price-inelastic; provides a small but consistent structural floor."),
    ("Structural Demand", "Bar & Coin Demand", "tonnes/quarter (t/qtr)",
     "<b>Frequency:</b> Quarterly (~6–8 weeks after quarter end). "
     "<b>Source:</b> <a href='https://www.gold.org/goldhub/research/gold-demand-trends' style='color:#7abfff'>WGC: Gold Demand Trends</a> "
     "(GDT_Tables_Q?xx_EN.xlsx, Gold Balance sheet row 20 — free registration required).<br>"
     "WGC quarterly retail investment demand (physical bullion, sovereign coins). "
     "Proxies for retail safe-haven sentiment and store-of-value buying, "
     "especially in Asia and the Middle East. Interpolated from quarterly to weekly."),
]

_method_html_parts = []
_cur_grp = "__UNSET__"
for r in method_rows:
    grp, name, unit, desc = r
    if grp is not None and grp != _cur_grp:
        _cur_grp = grp
        _method_html_parts.append(
            f'<tr style="background:#12141e">'
            f'<td colspan="3" style="padding:5px 10px;color:#c9a84c;font-size:10px;'
            f'font-weight:700;letter-spacing:0.7px;text-transform:uppercase">{grp}</td></tr>'
        )
    elif grp is None:
        _cur_grp = None
    _pl = "12px" if grp else "8px"
    _method_html_parts.append(
        f'<tr><td style="width:16%;vertical-align:top;padding-left:{_pl}"><b>{name}</b></td>'
        f'<td style="width:15%;text-align:center;vertical-align:top;color:#c9a84c;'
        f'font-family:monospace;font-size:11px;padding-top:9px">{unit}</td>'
        f'<td style="color:#aaa;font-size:12px;line-height:1.55">{desc}</td></tr>'
    )
method_html_rows = "\n".join(_method_html_parts)

# ── Forecast Log ───────────────────────────────────────────────────────────
LOG_CSV   = Path(__file__).parent / "data" / "forecast_log.csv"
_LOG_COLS = ["Date", "WindowStart", "WindowEnd",
             "FcstLow", "FcstMedian", "FcstHigh",
             "ActLow", "ActClose", "ActHigh", "Diff"]

def _add_months(d, n):
    m   = d.month - 1 + n
    yr  = d.year + m // 12
    mo  = m % 12 + 1
    day = min(d.day, _cal.monthrange(yr, mo)[1])
    return _date(yr, mo, day)

def _compute_fcst_range():
    """ANALYST-adjusted 3m model price: returns (low, median, high) at ±2 SD."""
    delta = p2_3m_w.get("intercept", 0)
    for col, info in fmeta.items():
        coef  = (p2_3m_w.get("factors", {}).get(col) or {}).get("coef", 0.0)
        rk    = info["raw_col"]
        fval  = ANALYST.get(rk, cur.get(rk))
        if fval is None:
            continue
        c     = cur.get(rk) or 0
        itype = info["input_type"]
        if itype == "diff_log":
            if fval > 0 and c > 0:
                delta += coef * math.log(fval / c)
        elif itype == "diff_level":
            delta += coef * (fval - c)
        else:
            delta += coef * fval
    ln3m = cur["ln_Gold_Nominal"] + delta
    return (round(math.exp(ln3m - 2 * rmse3m)),
            round(math.exp(ln3m)),
            round(math.exp(ln3m + 2 * rmse3m)))

def _load_log():
    if not LOG_CSV.exists():
        return []
    with open(LOG_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:                 # back-compat: fill missing FcstMedian for older rows
        r.setdefault("FcstMedian", "")
    return rows

def _save_log(rows):
    with open(LOG_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_LOG_COLS)
        w.writeheader()
        w.writerows(rows)

def _fetch_actual(win_start, win_end):
    """Return (low, close_on_end_date, high) for GC=F during window, or (None,None,None)."""
    try:
        import yfinance as yf
        from datetime import datetime as _dtt
        fetch_end = (_dtt.strptime(win_end, "%Y-%m-%d") + _td(days=5)).strftime("%Y-%m-%d")
        df = yf.download("GC=F", start=win_start, end=fetch_end,
                         progress=False, auto_adjust=True)
        if df.empty:
            return None, None, None
        if hasattr(df.columns, "levels"):          # flatten MultiIndex
            df.columns = df.columns.get_level_values(0)
        df = df[df.index.strftime("%Y-%m-%d") <= win_end]
        if df.empty:
            return None, None, None
        return (round(float(df["Low"].min())),
                round(float(df["Close"].iloc[-1])),
                round(float(df["High"].max())))
    except Exception as _e:
        print(f"  [forecast log] fetch_actual error: {_e}")
        return None, None, None

# ── Run log update ─────────────────────────────────────────────────────────
_today      = _date.today()
_today_str  = _today.strftime("%Y-%m-%d")
_win_start  = (_today + _td(days=1)).strftime("%Y-%m-%d")
_win_end    = _add_months(_today, 3).strftime("%Y-%m-%d")
_fcst_lo, _fcst_mid, _fcst_hi = _compute_fcst_range()

_log_rows = _load_log()

# Fill actuals for any completed window not yet filled
for _lr in _log_rows:
    if _lr.get("ActClose"):
        continue
    try:
        _wend_d = _date.fromisoformat(_lr["WindowEnd"])
    except Exception:
        continue
    if _wend_d <= _today:
        _a_lo, _a_cls, _a_hi = _fetch_actual(_lr["WindowStart"], _lr["WindowEnd"])
        if _a_cls is not None:
            _lr["ActLow"]   = _a_lo
            _lr["ActClose"] = _a_cls
            _lr["ActHigh"]  = _a_hi
            try:
                _ref = int(_lr["FcstMedian"]) if _lr.get("FcstMedian") else (int(_lr["FcstLow"]) + int(_lr["FcstHigh"])) // 2
                _lr["Diff"] = f"{(_a_cls - _ref) / _ref * 100:+.1f}%"
            except Exception:
                pass

# Today's entry — overwrite if already present, else append
_new_row = {"Date": _today_str, "WindowStart": _win_start, "WindowEnd": _win_end,
            "FcstLow": _fcst_lo, "FcstMedian": _fcst_mid, "FcstHigh": _fcst_hi,
            "ActLow": "", "ActClose": "", "ActHigh": "", "Diff": ""}
_existing_idx = next((i for i, r in enumerate(_log_rows)
                      if r.get("Date") == _today_str), None)
if _existing_idx is not None:
    _log_rows[_existing_idx] = _new_row
else:
    _log_rows.append(_new_row)

_save_log(_log_rows)

# ── Build log table HTML ───────────────────────────────────────────────────
def _fmtd(ds):
    try:
        from datetime import datetime as _dtt
        d = _dtt.strptime(ds, "%Y-%m-%d")
        return f"{d.day} {d.strftime('%b')} {d.year}"
    except Exception:
        return ds

def _fmtw(s, e):
    try:
        from datetime import datetime as _dtt
        sd = _dtt.strptime(s, "%Y-%m-%d")
        ed = _dtt.strptime(e, "%Y-%m-%d")
        return f"{sd.day} {sd.strftime('%b')} – {ed.day} {ed.strftime('%b')} {ed.year}"
    except Exception:
        return f"{s} – {e}"

def _log_row_html(r):
    has_act = bool(r.get("ActClose"))
    diff_s  = r.get("Diff", "")
    try:
        diff_c = "#4caf82" if float(diff_s.replace("%","")) >= 0 else "#e05c5c"
    except Exception:
        diff_c = "#888"
    # Colour actual close: green if within forecast band, red if outside
    act_cls_c = "#888"
    if has_act:
        try:
            _ac = int(r["ActClose"]); _fl = int(r["FcstLow"]); _fh = int(r["FcstHigh"])
            act_cls_c = "#4caf82" if _fl <= _ac <= _fh else "#e05c5c"
        except Exception:
            act_cls_c = "#b8bcc8"

    def _c(v, col):
        if v is None or v == "":
            return f'<td style="text-align:right;color:#3a3d4a;padding-right:10px">—</td>'
        try:
            return f'<td style="text-align:right;color:{col};padding-right:10px">${int(v):,}</td>'
        except Exception:
            return f'<td style="text-align:right;color:{col};padding-right:10px">{v}</td>'

    return (
        f'<tr>'
        f'<td style="white-space:nowrap">{_fmtd(r["Date"])}</td>'
        f'<td style="text-align:center;font-size:11px;color:#888">{_fmtw(r["WindowStart"], r["WindowEnd"])}</td>'
        f'{_c(r.get("FcstLow"),    "#7abfff")}'
        f'{_c(r.get("FcstMedian"), "#c9a84c")}'
        f'{_c(r.get("FcstHigh"),   "#7abfff")}'
        f'{_c(r.get("ActLow"),   "#ff9f4a" if has_act else "")}'
        f'{_c(r.get("ActClose"), act_cls_c)}'
        f'{_c(r.get("ActHigh"),  "#ff9f4a" if has_act else "")}'
        f'<td style="text-align:center;color:{diff_c if diff_s else "#3a3d4a"}">'
        f'{diff_s if diff_s else "—"}</td>'
        f'</tr>'
    )

_log_table_rows = ("\n".join(_log_row_html(r) for r in reversed(_log_rows))
                   if _log_rows else
                   '<tr><td colspan="8" style="text-align:center;color:#555;'
                   'padding:20px">No entries yet — run weekly to build the log.</td></tr>')

# ── HTML ───────────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Gold Price Estimator</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;
      background:#0f1117;color:#b8bcc8;font-size:14px;line-height:1.5}}
.container{{max-width:1100px;margin:0 auto;padding:24px 16px}}
h1{{font-size:22px;color:#c9a84c;font-weight:700;margin-bottom:4px}}
.subtitle{{color:#666;font-size:12px;margin-bottom:24px}}
h2{{font-size:16px;color:#c9a84c;font-weight:600;margin-bottom:10px;
    border-bottom:1px solid #2a2d3a;padding-bottom:6px}}
.card{{background:#1a1d27;border-radius:10px;padding:20px;
       border:1px solid #2a2d3a;margin-bottom:24px}}
.note{{color:#555;font-size:11px;margin-top:10px;font-style:italic}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#12141e;color:#c9a84c;font-weight:600;padding:9px 8px;
    border-bottom:2px solid #2a2d3a;white-space:nowrap;text-align:left}}
td{{padding:8px;border-bottom:1px solid #1f2235;vertical-align:middle}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#1f2235}}
.p1{{color:#7abfff}}.p2{{color:#ff9f4a}}
.gold-price{{font-size:42px;font-weight:700;color:#c9a84c}}
.gold-chg{{font-size:16px;margin-top:4px}}
.up{{color:#4caf82}}.down{{color:#e05c5c}}
input::-webkit-outer-spin-button,input::-webkit-inner-spin-button{{-webkit-appearance:none}}
details summary{{cursor:pointer;color:#c9a84c;font-size:13px;font-weight:600;
                 padding:4px 0;user-select:none}}
details[open] summary{{margin-bottom:12px}}
.sbox{{background:#12141e;border-radius:8px;padding:14px 18px;border:1px solid #2a2d3a}}
.sbox-label{{color:#888;font-size:11px;text-transform:uppercase;letter-spacing:0.4px}}
.sbox-val{{font-size:24px;font-weight:700;color:#c9a84c;margin-top:4px}}
.est-scroll{{max-height:100vh;overflow-y:auto;border-radius:0 0 8px 8px}}
.est-scroll table{{border-collapse:collapse}}
.est-scroll thead{{position:sticky;top:0;z-index:20;background:#1a1d27}}
.price-hdr td{{background:#12141e;border-bottom:2px solid #c9a84c33;padding:0}}
.info-bar{{background:#12141e;border-radius:8px;padding:10px 18px;margin-bottom:16px;
           display:flex;gap:24px;align-items:center;flex-wrap:wrap;border:1px solid #2a2d3a}}
.info-bar-item{{display:flex;flex-direction:column}}
.info-bar-label{{color:#888;font-size:10px;text-transform:uppercase;letter-spacing:0.3px}}
.info-bar-val{{font-size:15px;font-weight:700;color:#c9a84c;margin-top:1px}}
.r2-pill{{background:#0f1117;border-radius:6px;padding:5px 10px;font-size:12px}}
</style>
</head>
<body>
<div class="container">

<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:12px">
  <h1>Gold Price Estimator</h1>
  <span style="color:#888;font-size:12px">Last updated: {gen_display}</span>
</div>

<!-- ══ GOLD PRICE ESTIMATOR ══════════════════════════════════════════════ -->
<div class="card" style="padding:0;overflow:hidden">
  <div class="est-scroll">
    <table>
      <colgroup>
        <col style="width:18%"><col style="width:8%">
        <col style="width:7%"><col style="width:7%">
        <col style="width:7%"><col style="width:7%">
        <col style="width:11%"><col style="width:9%">
      </colgroup>
      <thead>
        <!-- Header row 1: Current gold + model signal (static) -->
        <tr class="price-hdr">
          <td colspan="8">
            <div style="display:flex;align-items:center;gap:0;flex-wrap:nowrap;font-size:12px;padding:0 4px">
              <div style="padding:6px 16px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">Current Gold</span>
                <span style="color:#c9a84c;font-size:20px;font-weight:700">${cur["Gold_Nominal"]:,.0f}</span>
                <span style="color:#888;font-size:11px;margin-left:6px">{cur["date"]}</span>
              </div>
              <div style="padding:6px 16px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">Past-3m Residual · z={fv_z:+.2f}</span>
                <span style="color:{fv_color(fv_pct3)};font-size:15px;font-weight:700">{fv_pct3:+.1f}%</span>
                <span style="color:#888;font-size:11px;margin-left:4px">gold {'above' if fv_pct3>0 else 'below'} model</span>
              </div>
              <div style="padding:6px 16px;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">Model Fair Value
                  <span style="color:#666;font-weight:normal;font-size:9px;margin-left:4px" title="Log-arithmetic: if gold is X% below model, it must gain X/(1-X%) to recover — slightly more than X%">why {fv_from_cur:+.1f}% &#8800; {fv_pct3:+.1f}%?</span>
                </span>
                <span style="color:#c9a84c;font-size:15px;font-weight:700">${f"{fv_price:,.0f}" if fv_price else "—"}</span>
                <span style="color:#888;font-size:11px;margin-left:6px">{fv_from_cur:+.1f}% upside · 6m: {fv_pct6:+.1f}%</span>
              </div>
            </div>
          </td>
        </tr>
        <!-- Header row 2: model forecast (locked) + revised forecast (live) -->
        <tr class="price-hdr">
          <td colspan="8">
            <div style="display:flex;align-items:center;gap:0;flex-wrap:nowrap;font-size:12px;padding:0 4px">
              <!-- Model forecast (set once at page load, never changes) -->
              <div style="padding:6px 16px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">3m Model Forecast</span>
                <span style="color:#7abfff;font-size:19px;font-weight:700" id="hdr-model-price">$—</span>
                <span style="font-size:11px;color:#7abfff;margin-left:6px" id="hdr-model-chg">—</span>
              </div>
              <div style="padding:6px 12px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">95% CI</span>
                <span style="color:#b8bcc8;font-size:12px"><span id="hdr-lo">—</span> – <span id="hdr-hi">—</span></span>
              </div>
              <div style="padding:6px 12px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">+ Mean Reversion</span>
                <span style="color:#7abfff;font-size:14px;font-weight:600" id="hdr-mr">$—</span>
                <span style="font-size:11px;color:#888;margin-left:4px" id="hdr-mr-chg">—</span>
              </div>
              <!-- Divider -->
              <div style="width:1px;background:#c9a84c44;align-self:stretch;margin:4px 6px"></div>
              <!-- Revised forecast (live — updates on input change) -->
              <div style="padding:6px 16px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#a0d8a0;font-size:10px;display:block;text-transform:uppercase">Revised Forecast</span>
                <span style="color:#a0d8a0;font-size:19px;font-weight:700" id="hdr-price">$—</span>
                <span style="font-size:11px;color:#a0d8a0;margin-left:6px" id="hdr-chg">—</span>
              </div>
              <div style="padding:6px 12px;border-right:1px solid #2a2d3a;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">Revised CI</span>
                <span style="color:#b8bcc8;font-size:12px"><span id="hdr-rev-lo">—</span> – <span id="hdr-rev-hi">—</span></span>
              </div>
              <div style="padding:6px 12px;white-space:nowrap">
                <span style="color:#888;font-size:10px;display:block;text-transform:uppercase">vs FV (${f"{fv_price:,.0f}" if fv_price else "—"})</span>
                <span style="font-size:13px;font-weight:700" id="hdr-fv-pct">—</span>
              </div>
            </div>
          </td>
        </tr>
        <!-- Column headers row 1 -->
        <tr>
          <th rowspan="2" style="cursor:pointer;padding:7px 8px" onclick="restoreFA()" title="Restore original order">Factor ↺</th>
          <th rowspan="2" style="text-align:center;padding:7px 4px">Current</th>
          <th colspan="2" class="p1" style="text-align:center;padding:7px 4px">QE Era</th>
          <th colspan="2" class="p2" style="text-align:center;padding:7px 4px">Dollar Era</th>
          <th rowspan="2" style="text-align:center;padding:7px 4px">3m Forecast<br><small style="color:#888;font-weight:normal">(editable)</small></th>
          <th rowspan="2" style="text-align:center;padding:7px 4px">Impact</th>
        </tr>
        <!-- Column headers row 2 -->
        <tr>
          <th class="p1" style="text-align:center;cursor:pointer;padding:5px 4px" onclick="sortFA('beta1')">β* ↕</th>
          <th class="p1" style="text-align:center;cursor:pointer;padding:5px 4px" onclick="sortFA('rp1')">Wt% ↕</th>
          <th class="p2" style="text-align:center;cursor:pointer;padding:5px 4px" onclick="sortFA('beta2')">β* ↕</th>
          <th class="p2" style="text-align:center;cursor:pointer;padding:5px 4px" onclick="sortFA('rp2')">Wt% ↕</th>
        </tr>
      </thead>
      <tbody id="fa-tbody">
{combined_html}
      </tbody>
    </table>
  </div>
</div>

<!-- ══ FORECAST LOG ══════════════════════════════════════════════════════ -->
<div class="card">
  <h2>&#128200; 3-Month Forecast Log</h2>
  <div style="overflow-x:auto">
    <table style="min-width:760px">
      <colgroup>
        <col style="width:10%"><col style="width:18%">
        <col style="width:8%"><col style="width:9%"><col style="width:8%">
        <col style="width:8%"><col style="width:9%"><col style="width:8%">
        <col style="width:7%">
      </colgroup>
      <thead>
        <tr>
          <th rowspan="2">Date Made</th>
          <th rowspan="2" style="text-align:center">Window (3m)</th>
          <th colspan="3" style="text-align:center;color:#7abfff">Forecast ±2 SD</th>
          <th colspan="3" style="text-align:center;color:#ff9f4a">Actual (window range)</th>
          <th rowspan="2" style="text-align:center">Diff</th>
        </tr>
        <tr>
          <th style="text-align:right;color:#7abfff;padding-right:10px">Low</th>
          <th style="text-align:right;color:#c9a84c;padding-right:10px">Price</th>
          <th style="text-align:right;color:#7abfff;padding-right:10px">High</th>
          <th style="text-align:right;color:#ff9f4a;padding-right:10px">Low</th>
          <th style="text-align:right;color:#ff9f4a;padding-right:10px">Close</th>
          <th style="text-align:right;color:#ff9f4a;padding-right:10px">High</th>
        </tr>
      </thead>
      <tbody>
{_log_table_rows}
      </tbody>
    </table>
  </div>
  <p class="note">Forecast Price = ANALYST-adjusted 3m model estimate; Low/High = Price ±2 SD (≈95% CI), captured at run time.
  Actual = GC=F during window: Low/High = period extremes, Close = last trading day of window.
  Diff = (actual close − forecast price) / forecast price. Close highlighted green if within forecast band.</p>
</div>

<!-- ══ METHODOLOGY + FOOTNOTES ═══════════════════════════════════════════ -->
<div class="card">
  <details>
    <summary>&#9658; Data &amp; Model Details</summary>
    <div style="margin-top:14px">
      <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px">
        <div class="r2-pill" style="border:1px solid #c9a84c33">
          <span style="color:#c9a84c;font-weight:600">3m weekly HAC ★ (primary)</span><br>
          <span style="color:#7abfff">{f"{p1_3m_w_r2:.3f} ({p1_3m_w_n}w)" if p1_3m_w_r2 else "—"}</span> QE &nbsp;/&nbsp;
          <span style="color:#ff9f4a">{f"{p2_3m_w_r2:.3f} ({p2_3m_w_n}w)" if p2_3m_w_r2 else "—"}</span> Dollar Era
        </div>
        <div class="r2-pill">
          <span style="color:#888">Monthly OLS</span><br>
          <span style="color:#7abfff">{p1["r2"]:.3f}</span> QE &nbsp;/&nbsp;
          <span style="color:#ff9f4a">{p2["r2"]:.3f}</span> Dollar
        </div>
        <div class="r2-pill">
          <span style="color:#888">Daily OLS</span><br>
          <span style="color:#7abfff">{f"{p1_daily_r2:.3f}" if p1_daily_r2 else "—"}</span> QE &nbsp;/&nbsp;
          <span style="color:#ff9f4a">{f"{p2_daily_r2:.3f}" if p2_daily_r2 else "—"}</span> Dollar
        </div>
      </div>
      <p style="color:#888;font-size:11px;margin-bottom:10px">
        <b>Model:</b> Δln(Gold_Nominal) OLS · Dollar Era split Mar 2022 (Russia reserve freeze) ·
        3m weekly HAC uses overlapping 13-week windows, Newey-West SE (lag=13) ·
        β* = coef × SD(X)/SD(Y) · Weight% = normalized partial R², sums to 100% ·
        1 SD (3m) ≈ {(math.exp(rmse3m)-1)*100:.1f}% ≈ ±${sd_dol:,} · 95% CI ≈ ±${ci_dol:,} ·
        CI covers full {f"{(1-p2_3m_w_r2)*100:.0f}%" if p2_3m_w_r2 else "~40%"} unexplained variance ·
        CB QE column uses Jan 2016–Feb 2022 monthly model ·
        Structural demand VIF elevated (WGC series correlated) — individual weights indicative ·
        Not financial advice.
      </p>
      <table style="margin-top:10px">
        <thead><tr>
          <th style="width:16%">Factor</th>
          <th style="width:16%;text-align:center">In-Model Form</th>
          <th>Purpose</th>
        </tr></thead>
        <tbody>{method_html_rows}</tbody>
      </table>
      <p style="margin:14px 0 2px;font-size:11px;color:#888;font-weight:600">Factors tested but not included:</p>
      <ul style="margin:4px 0 4px 18px;padding:0;color:#666;font-size:11px;line-height:1.75">
        <li><b>Geopolitical Risk Index (GPR)</b> (Caldara &amp; Iacoviello; daily, 1985–present;
          <a href='https://www.policyuncertainty.com/gpr.html' style='color:#7abfff'>policyuncertainty.com/gpr.html</a>) —
          carries the correct positive sign but adds no incremental explanatory power once TIPS, DXY, and ETF sentiment
          are controlled for (partial R² &lt;1% in both eras, n.s.). Excluded: parsimony, no incremental R².</li>
        <li><b>Bitcoin (BTC-USD log-return)</b> (daily;
          <a href='https://finance.yahoo.com/quote/BTC-USD/' style='color:#7abfff'>Yahoo Finance: BTC-USD</a>) —
          tested as a digital-gold substitution proxy; marginally negative sign in QE Era, significance disappears
          entirely in Dollar Era where TIPS and ETF already absorb speculative positioning. Excluded: parsimony, out-of-sample instability.</li>
      </ul>
    </div>
  </details>
</div>

</div>
<script>
const MODEL = {js_json};

function fmt(n) {{ return Math.round(n).toLocaleString("en-US"); }}

// ── Factor Analysis table sorting ─────────────────────────────────────────
let _faOrigHTML = "";
let _sortCol    = null;
let _sortDesc   = true;

function initFaSort() {{
  const tb = document.getElementById("fa-tbody");
  if (tb) _faOrigHTML = tb.innerHTML;
}}

function restoreFA() {{
  const tb = document.getElementById("fa-tbody");
  if (tb && _faOrigHTML) {{ tb.innerHTML = _faOrigHTML; _sortCol = null; }}
}}

function sortFA(col) {{
  const tb = document.getElementById("fa-tbody");
  if (!tb) return;

  if (_sortCol === col) {{ _sortDesc = !_sortDesc; }}
  else                  {{ _sortCol = col; _sortDesc = true; }}  // first click = desc (biggest first)

  const rows    = Array.from(tb.querySelectorAll("tr"));
  const sortable = rows.filter(r => r.classList.contains("fa-row"));
  const footer   = rows.filter(r => r.classList.contains("fa-footer"));

  sortable.sort((a, b) => {{
    const va = Math.abs(parseFloat(a.getAttribute("data-" + col)) || 0);
    const vb = Math.abs(parseFloat(b.getAttribute("data-" + col)) || 0);
    return _sortDesc ? vb - va : va - vb;
  }});

  tb.innerHTML = "";
  const label = col.startsWith("rp")   ? "Weight%" : "|β*|";
  const era   = col.endsWith("2")      ? "Dollar Era" : "QE Era";
  const dir   = _sortDesc ? "↓ highest" : "↑ lowest";
  const hdr   = document.createElement("tr");
  hdr.className = "fa-group-hdr";
  hdr.innerHTML = `<td colspan="8" style="background:#12141e;padding:10px 8px 6px;color:#c9a84c;font-weight:700;font-size:12px;letter-spacing:0.5px;text-transform:uppercase">` +
    `All Factors — sorted by ${{era}} ${{label}} ${{dir}} &nbsp;` +
    `<span style="color:#666;font-weight:normal;font-size:11px;cursor:pointer" onclick="restoreFA()">↺ restore</span></td>`;
  tb.appendChild(hdr);
  sortable.forEach(r => tb.appendChild(r));
  footer.forEach(r => tb.appendChild(r));
}}

function toggleEl(id) {{
  const el = document.getElementById(id);
  if (el) el.style.display = (el.style.display === "none" || el.style.display === "") ? "block" : "none";
}}

function onInput(el) {{
  const c   = parseFloat(el.dataset.current);
  const v   = parseFloat(el.value);
  const thr = (c > 100) ? 0.5 : 0.005;
  const chg = Math.abs(v - c) > thr;
  el.style.background  = chg ? "#2a1e00" : "#0f1117";
  el.style.borderColor = chg ? "#c9a84c66" : "#333";
  updateForecast();
}}

function updateTipsFromParts() {{
  const nom = parseFloat(document.getElementById("fi_nom_rate").value);
  const inf = parseFloat(document.getElementById("fi_inf_exp").value);
  if (!isNaN(nom) && !isNaN(inf)) {{
    const computed = (nom - inf).toFixed(2);
    document.getElementById("tips_computed_val").textContent = computed + "%";
    const tipsEl = document.getElementById("fi_D_TIPS_10yr");
    if (tipsEl) {{ tipsEl.value = computed; onInput(tipsEl); }}
  }}
}}

function updateStructTotal() {{
  const ids = ["fi_CB_Net_Purchases","fi_ETF_Residual","fi_Jewellery","fi_Technology","fi_Bar_Coin"];
  let total = 0;
  ids.forEach(id => {{
    const el = document.getElementById(id);
    if (el) {{ const v = parseFloat(el.value); if (!isNaN(v)) total += v; }}
  }});
  const disp = document.getElementById("fi_struct_total");
  if (disp) disp.textContent = Math.round(total).toLocaleString("en-US") + "t";
  updateForecast();
}}

function updateForecast() {{
  // 3m weekly HAC model. Orthogonalization (Stage 1a/1b) is already embedded in β coefficients.
  // Level  (M2, CB, ETF_Residual, Jewellery …): contrib = coef × fval;  pct = exp(coef×(fval−cur))−1
  // diff_log (DXY, Oil_Residual, VIX …):         contrib = coef × ln(fval/cur)
  // diff_level (TIPS, Credit):                    contrib = coef × (fval−cur)
  let delta = MODEL.intercept;

  for (const [col, info] of Object.entries(MODEL.factors)) {{
    const el  = document.getElementById("fi_" + col);
    const imp = document.getElementById("fi_impact_" + col);
    if (!el) continue;
    const fval = parseFloat(el.value);
    if (isNaN(fval)) {{ if (imp) imp.textContent = "—"; continue; }}

    const c = info.current;
    let contrib = 0, pct = 0;

    if (info.input_type === "diff_log") {{
      if (fval <= 0 || c <= 0) {{ if (imp) imp.textContent = "—"; continue; }}
      contrib = info.coef * (Math.log(fval) - Math.log(c));
      pct     = (Math.exp(contrib) - 1) * 100;
    }} else if (info.input_type === "diff_level") {{
      contrib = info.coef * (fval - c);
      pct     = (Math.exp(contrib) - 1) * 100;
    }} else {{
      // Level factor: 3m weekly coef × forecast level (quarterly scale, no ×3)
      contrib = info.coef * fval;
      pct     = (Math.exp(info.coef * (fval - c)) - 1) * 100;
    }}

    delta += contrib;
    if (imp) {{
      const s = pct >= 0 ? "+" : "", k = pct >= 0 ? "up" : "down";
      imp.innerHTML = `<span class="${{k}}">${{s}}${{pct.toFixed(2)}}%</span>`;
    }}
  }}

  // 3m gold forecast — RMSE already 3m scale (no √3)
  const ln3m = MODEL.currentLnGoldNom + delta;
  const g3m  = Math.exp(ln3m);
  const hci  = 1.96 * MODEL.rmse;
  const lo   = Math.exp(ln3m - hci);
  const hi   = Math.exp(ln3m + hci);
  const sd_d = Math.round(g3m * (Math.exp(MODEL.rmse) - 1));
  const ci_d = Math.round(g3m * (Math.exp(hci) - 1));
  const chg  = (g3m / MODEL.currentGoldNom - 1) * 100;
  const ks = chg >= 0 ? "+" : "", kk = chg >= 0 ? "up" : "down";

  // ── Forecast vs fair value
  const fvChg = MODEL.fvPrice ? (g3m / MODEL.fvPrice - 1) * 100 : null;
  const fvCol = fvChg === null ? "#888" : (fvChg > 5 ? "#e05c5c" : fvChg < -5 ? "#4caf82" : "#888");
  const fvS   = fvChg !== null ? (fvChg >= 0 ? "+" : "") + fvChg.toFixed(1) + "%" : "—";

  // ── Mean-reversion scenario (factors + full gap closure)
  let mrTxt = "—", mrChgTxt = "";
  if (MODEL.fvPrice) {{
    const g3m_mr = g3m * MODEL.fvPrice / MODEL.currentGoldNom;
    const chg_mr = (g3m_mr / MODEL.currentGoldNom - 1) * 100;
    mrTxt    = "$" + fmt(g3m_mr);
    mrChgTxt = (chg_mr >= 0 ? "+" : "") + chg_mr.toFixed(1) + "%";
  }}

  // ── Update frozen price header rows
  const _s = (id, txt) => {{ const e = document.getElementById(id); if (e) e.textContent = txt; }};
  const _h = (id, htm) => {{ const e = document.getElementById(id); if (e) e.innerHTML = htm; }};
  const _c = (id, col) => {{ const e = document.getElementById(id); if (e) e.style.color = col; }};
  const chgSpan = `<span class="${{kk}}">${{ks}}${{chg.toFixed(1)}}%</span>`;

  // Model forecast — locked after first call
  if (!window._modelFcstLocked) {{
    _s("hdr-model-price", "$" + fmt(g3m));
    _h("hdr-model-chg", chgSpan);
    _s("hdr-lo",     "$" + fmt(lo));
    _s("hdr-hi",     "$" + fmt(hi));
    _s("hdr-mr",     mrTxt);
    _s("hdr-mr-chg", mrChgTxt);
    window._modelFcstLocked = true;
  }}

  // Revised forecast — always live (green)
  _s("hdr-price",     "$" + fmt(g3m));
  _h("hdr-chg",      `<span style="color:#a0d8a0">${{ks}}${{chg.toFixed(1)}}%</span>`);
  _s("hdr-rev-lo",   "$" + fmt(lo));
  _s("hdr-rev-hi",   "$" + fmt(hi));
  _s("hdr-fv-pct",   fvS);
  _c("hdr-fv-pct",   fvCol);
}}

document.addEventListener("DOMContentLoaded", () => {{
  document.querySelectorAll('[id^="info_"]').forEach(el => el.style.display = "none");
  document.querySelectorAll('[id^="tips_"]').forEach(el => el.style.display = "none");
  initFaSort();  // capture original Factor Analysis table order for sort/restore
}});

updateStructTotal();  // initialises structural total and triggers updateForecast
</script>
</body>
</html>"""

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)

print(f"Dashboard -> {OUT_HTML}")

# ── Factor forecast table ──────────────────────────────────────────────────
# ── Factor forecast sources ───────────────────────────────────────────────
FORECAST_SOURCES = {
    "TIPS_10yr":        "FRED DFII10 Jun 26 2026 = 2.17% (current); 3m consensus ~2.04% (Trading Economics); use 2.10 midpoint",
    "DXY":              "Current 101.4; JPM Q3 103-104; Fed on hold (CPI >4%); no rate cuts priced; use 102.5",
    "M2_Growth":        "FRED M2SL trailing 12m growth trend; no analyst consensus — 5.5% is midpoint of 4.5-6% recent range",
    "VIX":              "No analyst consensus; historical long-run mean 2010-2024 = ~17-19; 18 = neutral regime",
    "EPU":              "No analyst consensus; using recent trailing level (~225); Baker-Bloom-Davis index",
    "Credit_Spread":    "BAA10Y recent trend ~1.51%; slight widening to 1.55% — no tight consensus",
    "QQQ":              "Goldman Sachs S&P 500 target raised to 8,000 (May 26 2026) from ~7,519 = +6.4%; applied to QQQ",
    "Oil_WTI":          "Iran war premium unwound; Strait of Hormuz recovering; EIA Q3 range $65-78/bbl; current ~$70; base case $72",
    "CB_Net_Purchases": "WGC Q1 2026 actual 244t; FY 850t guidance; Jun 2026 WGC survey: 45% of CBs plan to increase holdings",
    "ETF_Flow":         "WGC Apr 2026 +45t, May 2026 -16t; Q2 est ~30t (slowing from Q1 +62t); momentum cooling",
    "Breakeven":        "FRED T10YIE Jun 26 2026 = 2.17%; used for TIPS decompose display only",
    "CPI":              "REMOVED from model — TIPS already captures inflation; kept for reference only",
    "DJP":              "Neutral: energy drag from oil decline offset by metals/ag; keep at current level (no directional call)",
}

print()
print("=" * 90)
print("3-MONTH FACTOR FORECAST TABLE — Q3 2026 ANALYST CONSENSUS SOURCES")
print("OLS DV: Delta ln(Gold_Nominal) — nominal price directly, no CPI adjustment")
print("=" * 90)
print(f"  {'Factor':<28} {'Now':>8} {'Fcst':>8} {'Change':>13}  Source")
print("-" * 90)
for col, meta in fmeta.items():
    rk       = meta["raw_col"]
    cv       = cur.get(rk)
    fcast    = ANALYST.get(rk)
    itype    = meta["input_type"]
    source   = FORECAST_SOURCES.get(rk, "—")
    if cv is None or fcast is None:
        continue
    if itype == "diff_log":
        chg_str = f"{(math.log(fcast/cv))*100:+.2f}% log"
    elif itype == "diff_level":
        chg_str = f"{fcast - cv:+.3f} pp"
    else:
        chg_str = f"lev {fcast:.2f}"
    label = meta["label"]
    print(f"  {label:<28} {cv:>8.3f} {fcast:>8.3f} {chg_str:>13}  {source}")

print()
print("  TIPS Decompose (info only — not a separate OLS factor):")
_tips_now2  = cur.get("TIPS_10yr", 0)
_beven_now2 = cur.get("Breakeven",  0)
print(f"    Nominal rate now:   ~{_tips_now2+_beven_now2:.2f}%   Forecast: ~{_nom_fcast:.2f}%")
print(f"    Inflation expect:   ~{_beven_now2:.2f}%   Forecast: ~{_beven_fcast:.2f}%  "
      f"(Source: {FORECAST_SOURCES['Breakeven']})")
print(f"    TIPS (real):        ~{_tips_now2:.2f}%   Forecast: ~{_tips_fcast:.2f}%  --> THIS enters OLS")
print()
print("  NOTE: Impact on Gold = coeff x change")
print("        diff_level: change = forecast - current (in percentage points)")
print("        diff_log:   change = ln(forecast/current) ~ percent change")
print("        level:      monthly contrib = 3 x coeff x forecast_value (for M2/CB)")
